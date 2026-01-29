"""XLSX parser implementation."""

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from ..models import Chunk, ChunkType, Heading, Locator
from .base import Parser


class XLSXParser(Parser):
    """Parser for XLSX files."""

    def parse(self, file_path: Path, doc_id: str, company_id: str) -> tuple[list[Chunk], list[Heading]]:
        """Parse an XLSX file."""
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        chunks: list[Chunk] = []
        headings: list[Heading] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Create heading for sheet
            heading_id = hashlib.md5(f"{doc_id}:{sheet_name}".encode()).hexdigest()[:16]
            heading = Heading(
                heading_id=heading_id,
                doc_id=doc_id,
                company_id=company_id,
                heading_text=sheet_name,
                heading_norm=sheet_name.lower().replace(" ", "_"),
                level=1,
                locator=Locator(sheet=sheet_name),
            )
            headings.append(heading)

            # Extract data from sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                # Skip empty rows
                if any(cell.strip() for cell in row_data):
                    data.append(row_data)

            if not data:
                continue

            # Create chunk for the entire sheet
            text_lines = []
            for row in data:
                text_lines.append(" | ".join(row))
            
            text = "\n".join(text_lines)
            
            if text.strip():
                chunk_id = hashlib.md5(
                    f"{doc_id}:{sheet_name}:full".encode()
                ).hexdigest()[:16]
                
                chunk = Chunk(
                    chunk_id=chunk_id,
                    doc_id=doc_id,
                    company_id=company_id,
                    chunk_type=ChunkType.SHEET_RANGE,
                    text=text,
                    structure_path=[sheet_name],
                    locator=Locator(
                        sheet=sheet_name,
                        cell_range=f"A1:{sheet.max_column}{sheet.max_row}",
                    ),
                    raw_table=data,
                )
                chunks.append(chunk)

        wb.close()
        return chunks, headings
