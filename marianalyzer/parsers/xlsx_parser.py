"""XLSX document parser using openpyxl."""

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from marianalyzer.models import Chunk, Document, ParsedDocument
from marianalyzer.parsers.base import BaseParser
from marianalyzer.utils.citations import format_citation
from marianalyzer.utils.logging_config import get_logger

logger = get_logger()


class XLSXParser(BaseParser):
    """Parser for XLSX documents."""

    def supports_format(self, extension: str) -> bool:
        """Check if extension is .xlsx."""
        return extension.lower() == ".xlsx"

    def parse(self, file_path: Path) -> ParsedDocument:
        """Parse XLSX document.

        Args:
            file_path: Path to XLSX file

        Returns:
            ParsedDocument with chunks (one per row)
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        logger.info(f"Parsing XLSX: {file_path}")

        try:
            workbook = load_workbook(str(file_path), data_only=True)

            chunks = []
            chunk_index = 0

            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get headers from first row
                headers = []
                first_row = True

                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # Skip completely empty rows
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        continue

                    # First row is headers
                    if first_row:
                        headers = [str(cell) if cell is not None else f"Col{i}" for i, cell in enumerate(row, 1)]
                        first_row = False
                        continue

                    # Create row text with headers
                    row_data = []
                    min_cell_col = None
                    max_cell_col = None

                    for col_index, (header, value) in enumerate(zip(headers, row), start=1):
                        if value is not None and str(value).strip():
                            row_data.append(f"{header}: {value}")
                            if min_cell_col is None:
                                min_cell_col = col_index
                            max_cell_col = col_index

                    if not row_data:
                        continue

                    row_text = " | ".join(row_data)

                    # Create cell reference (e.g., "A5:C5")
                    if min_cell_col and max_cell_col:
                        start_cell = self._col_num_to_letter(min_cell_col) + str(row_index)
                        end_cell = self._col_num_to_letter(max_cell_col) + str(row_index)
                        cell_ref = f"{start_cell}:{end_cell}" if start_cell != end_cell else start_cell
                    else:
                        cell_ref = f"Row{row_index}"

                    citation = format_citation(
                        file_path=str(file_path.name),
                        sheet=sheet_name,
                        cell=cell_ref,
                    )

                    chunk = Chunk(
                        doc_id=0,
                        chunk_index=chunk_index,
                        chunk_text=row_text,
                        chunk_type="table_row",
                        citation=citation,
                        metadata={
                            "sheet": sheet_name,
                            "row": row_index,
                            "cell_ref": cell_ref,
                            "headers": headers,
                        },
                    )
                    chunks.append(chunk)
                    chunk_index += 1

            logger.info(f"Extracted {len(chunks)} chunks from {len(workbook.sheetnames)} sheets")

            # Create document metadata
            file_hash = self._compute_file_hash(file_path)

            document = Document(
                file_path=str(file_path.name),
                file_hash=file_hash,
                file_type="xlsx",
                file_size=file_path.stat().st_size,
                metadata={
                    "num_sheets": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                },
            )

            return ParsedDocument(
                metadata=document,
                chunks=chunks,
                headings=[],
            )

        except Exception as e:
            logger.error(f"Failed to parse XLSX {file_path}: {e}")
            raise

    def _col_num_to_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter (1 -> A, 27 -> AA, etc.)."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result

    def _compute_file_hash(self, file_path: Path) -> str:
        """Compute SHA256 hash of file."""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
